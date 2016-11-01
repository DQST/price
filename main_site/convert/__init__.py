import openpyxl
from openpyxl.utils import *
from yattag import *
import codecs


class BConverter(object):
	def __init__(self, path, head_row=0, start_row=1):
		wb = openpyxl.load_workbook(path, data_only=True)
		self.sh = wb.active
		self.headers = [str(i.value).strip() for i in self.sh[head_row:head_row] if i.value is not None]
		self.start_row = abs(start_row) - 1
		self.objects = self.__parse(head_row)

	def __unmerged_value(self, cell):
		for irange in self.sh.merged_cell_ranges:
			min_col, min_row, max_col, max_row =range_boundaries(irange)
			if cell.row in range(min_row,max_row+1) and column_index_from_string(cell.column) in range(min_col, max_col+1):
				return self.sh.cell(None, min_row, min_col).value
		return cell.value

	def __empty(self, o):
		e = 0
		for i in o:
			if i[1] == '':
				e += 1
		if e == len(o):
			return True
		return False

	def __parse(self, head_row):
		out = []
		fun = lambda x: '' if x is None else x
		for row in self.sh.iter_rows(row_offset=self.start_row):
			obj = [(key, fun(self.__unmerged_value(el))) for key, el in zip(self.headers, row)]
			if not self.__empty(obj):
				out.append(obj)
		return out


class ToXML(BConverter):
	def save(self, path):
		'''Doc string:
		Save file as xml using argument: path'''
		doc, tag, text, line = Doc().ttl()

		headers_len = len(self.headers)
		objects_len = len(self.objects)
		doc.asis('<?xml version="1.0" encoding="utf-8"?>')
		with tag('root'):
			with tag('headers', count=headers_len):
				for i, key in zip(range(1, headers_len + 1), self.headers):
					line('head', key, id=str(i))
			with tag('objects', count=objects_len):
				for i, obj in zip(range(1, objects_len + 1), self.objects):
					with tag('object', id=str(i)):
						for key, value in obj:
							with tag('field', name=key):
								text(value)

		pretty_string = indent(
			doc.getvalue(),
			indentation = '\t',
			newline = '\n',
		)

		with codecs.open(path, 'w', encoding='utf-8') as f:
			f.write('%s\n' % pretty_string)


class XLStoXLSX:
	def __init__(self, path):
		self.input_path = path
	
	def __unmerged_value(self, rowx, colx, thesheet):
		for crange in thesheet.merged_cells:
			rlo, rhi, clo, chi = crange
			if rowx in range(rlo, rhi):
				if colx in range(clo, chi):
					return thesheet.cell_value(rlo,clo)
		return thesheet.cell_value(rowx,colx)
	
	def save(self, path):
		import xlrd
		wb = xlrd.open_workbook(self.input_path, formatting_info=True)
		xlsx_wb = openpyxl.Workbook()
		xlsx_wb._sheets.clear()
		for name in wb.sheet_names():
			sh = wb.sheet_by_name(name)
			xlsx_wb.create_sheet(name)
			sh2 = xlsx_wb.active
			for i in range(sh.nrows):
				for j in range(sh.ncols):
					val = self.__unmerged_value(i, j, sh)
					sh2.cell(row=i+1, column=j+1, value=val)
		xlsx_wb.save(path)
